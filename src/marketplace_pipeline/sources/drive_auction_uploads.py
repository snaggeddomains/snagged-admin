"""Drive auction uploads — manually-uploaded auction CSVs/XLSX.

Port of legacy/openclaw/scripts/scan_drive_auction_uploads.py. Scans the
Drive folder where Judy / collaborators drop auction lists (typically
NameJet exports), parses each file, and publishes a consolidated row set
to the auctions sheet + #auctions Slack section.

This is the replacement for the legacy direct-NameJet sources (last
chance, email digest, exclusive storefront) — all three are disabled in
the registry; this single Drive-scan source handles whatever ends up in
the upload folder regardless of which NameJet feed produced it.

Uses the Google SA already configured. The legacy upload filter is
intentionally stricter than the standard SNAP filter:
  - TLDs limited to .com/.org/.io/.co
  - SLD must be <= 15 chars
  - No digits, no hyphens
"""
from __future__ import annotations

import csv
import io
import os
import re
from datetime import datetime, timedelta, timezone
from typing import Any

from .. import config, drive_cache, state
from ..auctions import sheet as auctions_sheet
from ..auctions import slack as auctions_slack
from ..filters import standard as flt

SOURCE_ID = "drive_auction_uploads"
SOURCE_LABEL = "Drive Uploads"
PLATFORM = "Drive Upload"

WINDOW_HOURS = 36
MAX_FILES = 20

# Stricter than the standard SNAP filter — uploads usually come from
# user-curated lists where junk should be aggressively rejected.
ALLOWED_UPLOAD_TLDS = (".com", ".org", ".io", ".co")
MAX_UPLOAD_SLD_LEN = 15

# Generic column-name fallbacks for non-NameJet CSV/XLSX
DOMAIN_COLS = ("domain", "domain name", "name")
PRICE_COLS = ("price", "buy now", "bin", "minimum bid", "min bid", "current bid", "bid")
END_COLS = (
    "order by", "close", "close date", "close time",
    "end", "end date", "end time", "auction end",
)
BID_COLS = ("bidders", "bidder count", "bids")
LINK_COLS = ("url", "link")

# Treat upload-listed times as Eastern (legacy parity)
EASTERN = timezone(timedelta(hours=-4))

SHEET_URL_TEMPLATE = "https://docs.google.com/spreadsheets/d/{sheet_id}/edit"
SNAPSHOT_FILE = "snapshot.json"


def upload_filter(domain: str) -> bool:
    """Apply the legacy upload-specific filter (stricter than allow_domain)."""
    domain = (domain or "").strip().lower()
    sld, tld = flt.extract_sld_tld(domain)
    if tld not in ALLOWED_UPLOAD_TLDS:
        return False
    if len(sld) > MAX_UPLOAD_SLD_LEN:
        return False
    if any(ch.isdigit() for ch in sld):
        return False
    if "-" in sld:
        return False
    if not flt.allow_domain(domain):
        return False
    return True


def _parse_money(value: Any) -> float | None:
    if value in (None, ""):
        return None
    cleaned = str(value).strip().replace("$", "").replace(",", "")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except (TypeError, ValueError):
        return None


def _parse_dt(value: Any) -> datetime | None:
    """Parse an upload-listed datetime. Many formats are used; try them in
    order. Returns a UTC datetime, or None if unparseable / 'available'."""
    if not value:
        return None
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    if not text or text.lower() in {"available", "available soon"}:
        return None
    for fmt in (
        "%b %d, %Y %I:%M:%S %p",
        "%b %d, %Y %I:%M %p",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y %I:%M %p",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    ):
        try:
            dt = datetime.strptime(text, fmt).replace(tzinfo=EASTERN)
            return dt.astimezone(timezone.utc)
        except (ValueError, TypeError):
            continue
    try:
        dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except (ValueError, TypeError):
        return None


def _normalize_headers(headers: list[Any]) -> list[str]:
    return [str(h or "").strip() for h in headers]


def _find_col(headers: list[str], candidates: tuple[str, ...]) -> str | None:
    normalized = {h.lower().strip(): h for h in headers if h}
    for c in candidates:
        if c in normalized:
            return normalized[c]
    return None


def _rows_from_csv(raw: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = raw.decode("utf-8-sig", "ignore")
    reader = csv.DictReader(io.StringIO(text))
    headers = _normalize_headers(reader.fieldnames or [])
    return headers, list(reader)


def _rows_from_xlsx(raw: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return [], []
    headers = _normalize_headers(list(values[0]))
    rows: list[dict[str, Any]] = []
    for row in values[1:]:
        item: dict[str, Any] = {}
        for idx, h in enumerate(headers):
            item[h] = row[idx] if idx < len(row) else ""
        rows.append(item)
    return headers, rows


def _parse_namejet_like(rows: list[dict[str, Any]], meta: dict[str, Any]) -> list[dict[str, Any]]:
    """NameJet-style upload: has 'Domain Name' + 'Order By' columns."""
    out: list[dict[str, Any]] = []
    for row in rows:
        domain = str(row.get("Domain Name") or "").strip().lower()
        if not upload_filter(domain):
            continue
        end_dt = _parse_dt(row.get("Order By"))
        out.append({
            "domain": domain,
            "platform": "NameJet Upload",
            "end_time_utc": end_dt.isoformat() if end_dt else None,
            "price": _parse_money(row.get("Minimum Bid")),
            "bid_count": _coerce_int(row.get("Bidders")),
            "link": f"https://www.namejet.com/domain/{domain}.action",
            "source_file": meta["name"],
            "source_file_modified": meta.get("modifiedTime"),
        })
    return out


def _coerce_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _parse_generic(
    rows: list[dict[str, Any]],
    headers: list[str],
    meta: dict[str, Any],
) -> list[dict[str, Any]]:
    """Generic upload — uses fuzzy column-name detection."""
    domain_col = _find_col(headers, DOMAIN_COLS)
    if not domain_col:
        return []
    price_col = _find_col(headers, PRICE_COLS)
    end_col = _find_col(headers, END_COLS)
    bid_col = _find_col(headers, BID_COLS)
    link_col = _find_col(headers, LINK_COLS)
    out: list[dict[str, Any]] = []
    for row in rows:
        domain = str(row.get(domain_col) or "").strip().lower()
        if not upload_filter(domain):
            continue
        end_dt = _parse_dt(row.get(end_col)) if end_col else None
        out.append({
            "domain": domain,
            "platform": PLATFORM,
            "end_time_utc": end_dt.isoformat() if end_dt else None,
            "price": _parse_money(row.get(price_col)) if price_col else None,
            "bid_count": _coerce_int(row.get(bid_col)) if bid_col else None,
            "link": (str(row.get(link_col) or "").strip() if link_col else None) or None,
            "source_file": meta["name"],
            "source_file_modified": meta.get("modifiedTime"),
        })
    return out


def detect_and_parse(meta: dict[str, Any], raw: bytes) -> list[dict[str, Any]]:
    """Dispatch on file type + header shape."""
    name = meta["name"]
    lower = name.lower()
    csv_mimes = {"text/csv", "application/csv"}
    xlsx_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if lower.endswith(".csv") or meta.get("mimeType") in csv_mimes:
        headers, rows = _rows_from_csv(raw)
    elif lower.endswith(".xlsx") or meta.get("mimeType") == xlsx_mime:
        headers, rows = _rows_from_xlsx(raw)
    else:
        return []
    if "Domain Name" in headers and "Order By" in headers:
        return _parse_namejet_like(rows, meta)
    return _parse_generic(rows, headers, meta)


def dedupe_listings(listings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """One row per domain — prefer the more recent source file."""
    best: dict[str, dict[str, Any]] = {}
    for row in listings:
        domain = row["domain"]
        prior = best.get(domain)
        if not prior:
            best[domain] = row
            continue
        if (row.get("source_file_modified") or "") > (prior.get("source_file_modified") or ""):
            best[domain] = row
            continue
        # Same modified time — prefer one with a price over one without
        if prior.get("price") is None and row.get("price") is not None:
            best[domain] = row
    return sorted(best.values(), key=lambda r: (r.get("end_time_utc") or "9999", r["domain"]))


def run() -> int:
    reg = config.load_registry()
    src_cfg = config.get_source(SOURCE_ID)
    auc_cfg = reg["products"]["auctions"]
    sheet_id = auc_cfg["sheet_id"]
    slack_channel = os.environ.get(auc_cfg["slack_channel_env"], "C096AT8BECS")
    sheet_url = SHEET_URL_TEMPLATE.format(sheet_id=sheet_id)
    folder_id = src_cfg["fetch"]["drive_folder_id"]
    window_hours = src_cfg["fetch"].get("window_hours", WINDOW_HOURS)

    print(f"[1/5] Listing files in Drive folder {folder_id}")
    files = drive_cache.list_files_in_folder(folder_id, page_size=MAX_FILES)
    cutoff = datetime.now(timezone.utc) - timedelta(hours=window_hours)
    recent = [
        f for f in files
        if f.get("modifiedTime") and datetime.fromisoformat(
            f["modifiedTime"].replace("Z", "+00:00")
        ) >= cutoff
    ]
    print(f"      {len(files):,} total, {len(recent):,} modified in last {window_hours}h")

    print("[2/5] Downloading + parsing each recent file")
    all_listings: list[dict[str, Any]] = []
    file_stats: list[dict[str, Any]] = []
    for meta in recent[:MAX_FILES]:
        try:
            raw = drive_cache.download_file(meta["id"])
        except Exception as e:
            print(f"      WARN failed to download {meta['name']}: {e}")
            continue
        try:
            listings = detect_and_parse(meta, raw)
        except Exception as e:
            print(f"      WARN failed to parse {meta['name']}: {e}")
            continue
        file_stats.append({
            "name": meta["name"],
            "modifiedTime": meta.get("modifiedTime"),
            "qualified_rows": len(listings),
        })
        all_listings.extend(listings)
        print(f"      {meta['name']}: kept {len(listings):,} rows")

    print(f"[3/5] Deduping (kept newer source-file per domain)")
    deduped = dedupe_listings(all_listings)
    # Drop entries without parseable end_time — auctions sheet requires it
    deduped = [L for L in deduped if L.get("end_time_utc")]
    print(f"      after dedup + end-time filter: {len(deduped):,}")

    now = datetime.now(timezone.utc)
    sheet_rows = [auctions_sheet.row_from_listing(L, now=now) for L in deduped]

    print("[4/5] Writing to auctions sheet")
    sheet_stats = auctions_sheet.write(
        spreadsheet_id=sheet_id,
        new_rows=sheet_rows,
    )
    print(f"      stats: {sheet_stats}")

    print("[5/5] Saving snapshot + Slack post")
    state.write_json(SOURCE_ID, SNAPSHOT_FILE, deduped)

    slack_listings = []
    for L in deduped:
        end_dt = datetime.fromisoformat(L["end_time_utc"].replace("Z", "+00:00"))
        slack_listings.append({
            **L,
            "time_left": auctions_sheet.format_time_left(end_dt, now=now),
        })

    section = auctions_slack.format_section(label=SOURCE_LABEL, listings=slack_listings)
    posted = auctions_slack.post_consolidated(
        channel=slack_channel,
        source=SOURCE_ID,
        sections=[section],
        sheet_url=sheet_url,
    )
    print(f"      slack posted: {posted}")

    state.write_json(SOURCE_ID, "run_status.json", {
        "source": SOURCE_ID,
        "label": SOURCE_LABEL,
        "status": "ok",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "files_scanned": len(recent),
        "rows_after_dedup": len(deduped),
        "new_count": sheet_stats["added"],
        "sheet_total_after": sheet_stats["total_after"],
        "deduped_against_existing": sheet_stats["deduped"],
        "slack_posted": posted,
        "files": file_stats,
    })

    print("DONE")
    return 0
