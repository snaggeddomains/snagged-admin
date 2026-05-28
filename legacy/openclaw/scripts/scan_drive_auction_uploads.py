#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
from datetime import datetime, timedelta, timezone
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable

import openpyxl
from google.oauth2 import service_account
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

from domain_filters import allow_domain
from namejet_digest_filter import DigestRow

BASE = Path('/root/.openclaw/workspace')
FOLDER_ID = '1vCnJb4iJeVJnLiRk4BwO7TEbRY16-Gta'
TOKEN = BASE / '.secrets/google_oauth_credentials.json'
SERVICE_ACCOUNT = BASE / '.secrets/google_service_account.json'
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
OUTPUT = BASE / 'data/drive_uploads_filtered.json'
RAW_DIR = BASE / 'data/drive_uploads/raw'
WINDOW_HOURS = 36
MAX_FILES = 20
DOMAIN_COLS = ('domain', 'domain name', 'name')
PRICE_COLS = ('price', 'buy now', 'bin', 'minimum bid', 'min bid', 'current bid', 'bid')
END_COLS = ('order by', 'close', 'close date', 'close time', 'end', 'end date', 'end time', 'auction end')
BID_COLS = ('bidders', 'bidder count', 'bids')
LINK_COLS = ('url', 'link')
ET = timezone(timedelta(hours=-4))
ALLOWED_UPLOAD_TLDS = {'.com', '.org', '.io', '.co'}
MAX_UPLOAD_SLD_LEN = 15


def load_drive():
    if SERVICE_ACCOUNT.exists():
        creds = service_account.Credentials.from_service_account_file(str(SERVICE_ACCOUNT), scopes=SCOPES)
        return build('drive', 'v3', credentials=creds, cache_discovery=False)
    creds = Credentials.from_authorized_user_file(str(TOKEN), scopes=SCOPES)
    return build('drive', 'v3', credentials=creds, cache_discovery=False)


def list_recent_files(service) -> list[dict]:
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=WINDOW_HOURS)).isoformat().replace('+00:00', 'Z')
    q = (
        f"'{FOLDER_ID}' in parents and trashed = false and modifiedTime >= '{cutoff}' "
        "and (mimeType != 'application/vnd.google-apps.folder')"
    )
    resp = service.files().list(
        q=q,
        orderBy='modifiedTime desc',
        pageSize=MAX_FILES,
        fields='files(id,name,mimeType,modifiedTime,webViewLink,size)',
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    return resp.get('files', [])


def download_file(service, meta: dict) -> bytes:
    request = service.files().get_media(fileId=meta['id'], supportsAllDrives=True)
    fh = BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return fh.getvalue()


def parse_money(value) -> float | None:
    if value in (None, ''):
        return None
    cleaned = str(value).strip().replace('$', '').replace(',', '')
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_dt(value: str | None) -> str | None:
    if not value:
        return None
    text = re.sub(r'\s+', ' ', str(value).replace('\xa0', ' ')).strip()
    if not text or text.lower() in {'available', 'available soon'}:
        return None
    fmts = [
        '%b %d, %Y %I:%M:%S %p',
        '%b %d, %Y %I:%M %p',
        '%m/%d/%Y %H:%M',
        '%m/%d/%Y %I:%M %p',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
    ]
    for fmt in fmts:
        try:
            dt = datetime.strptime(text, fmt).replace(tzinfo=ET)
            return dt.astimezone(timezone.utc).isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace('Z', '+00:00')).astimezone(timezone.utc).isoformat()
    except Exception:
        return None


def upload_filter(domain: str) -> bool:
    domain = (domain or '').strip().lower()
    if not allow_domain(domain, allowed_tlds=tuple(ALLOWED_UPLOAD_TLDS)):
        return False
    sld = domain.split('.', 1)[0]
    if len(sld) > MAX_UPLOAD_SLD_LEN:
        return False
    if any(ch.isdigit() for ch in sld) or '-' in sld:
        return False
    return True


def normalize_headers(headers: Iterable) -> list[str]:
    return [str(h or '').strip() for h in headers]


def find_col(headers: list[str], candidates: tuple[str, ...]) -> str | None:
    normalized = {h.lower().strip(): h for h in headers if h}
    for c in candidates:
        if c in normalized:
            return normalized[c]
    return None


def rows_from_csv_bytes(raw: bytes) -> tuple[list[str], list[dict]]:
    text = raw.decode('utf-8-sig', 'ignore')
    reader = csv.DictReader(StringIO(text))
    headers = normalize_headers(reader.fieldnames or [])
    return headers, list(reader)


def rows_from_xlsx_bytes(raw: bytes) -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return [], []
    headers = normalize_headers(values[0])
    rows = []
    for row in values[1:]:
        item = {headers[idx]: row[idx] if idx < len(row) else '' for idx in range(len(headers))}
        rows.append(item)
    return headers, rows


def parse_namejet_like(rows: list[dict], source_name: str, meta: dict) -> list[dict]:
    out = []
    for row in rows:
        domain = str(row.get('Domain Name') or '').strip().lower()
        if not upload_filter(domain):
            continue
        out.append({
            'domain': domain,
            'closing_dt_utc': parse_dt(row.get('Order By')),
            'min_bid_value': parse_money(row.get('Minimum Bid')),
            'bidders': str(row.get('Bidders') or '0').strip() or '0',
            'link': f'https://www.namejet.com/domain/{domain}.action',
            'platform': 'NameJet Upload',
            'source_file': source_name,
            'source_file_id': meta['id'],
            'source_file_modified': meta.get('modifiedTime'),
            'source_link': meta.get('webViewLink'),
        })
    return out


def parse_generic(rows: list[dict], headers: list[str], source_name: str, meta: dict) -> list[dict]:
    domain_col = find_col(headers, DOMAIN_COLS)
    if not domain_col:
        return []
    price_col = find_col(headers, PRICE_COLS)
    end_col = find_col(headers, END_COLS)
    bid_col = find_col(headers, BID_COLS)
    link_col = find_col(headers, LINK_COLS)
    out = []
    for row in rows:
        domain = str(row.get(domain_col) or '').strip().lower()
        if not upload_filter(domain):
            continue
        out.append({
            'domain': domain,
            'closing_dt_utc': parse_dt(row.get(end_col)) if end_col else None,
            'min_bid_value': parse_money(row.get(price_col)) if price_col else None,
            'bidders': str(row.get(bid_col) or '').strip() if bid_col else '',
            'link': str(row.get(link_col) or '').strip() if link_col else '',
            'platform': 'Drive Upload',
            'source_file': source_name,
            'source_file_id': meta['id'],
            'source_file_modified': meta.get('modifiedTime'),
            'source_link': meta.get('webViewLink'),
        })
    return out


def detect_and_parse(meta: dict, raw: bytes) -> list[dict]:
    name = meta['name']
    lower = name.lower()
    headers: list[str]
    rows: list[dict]
    if lower.endswith('.csv') or meta.get('mimeType') in {'text/csv', 'application/csv'}:
        headers, rows = rows_from_csv_bytes(raw)
    elif lower.endswith('.xlsx') or meta.get('mimeType') == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        headers, rows = rows_from_xlsx_bytes(raw)
    else:
        return []

    if 'Domain Name' in headers and 'Order By' in headers:
        return parse_namejet_like(rows, name, meta)
    return parse_generic(rows, headers, name, meta)


def dedupe(rows: list[dict]) -> list[dict]:
    best = {}
    for row in rows:
        domain = row['domain']
        prior = best.get(domain)
        if not prior:
            best[domain] = row
            continue
        if (row.get('source_file_modified') or '') > (prior.get('source_file_modified') or ''):
            best[domain] = row
            continue
        if (prior.get('min_bid_value') is None) and (row.get('min_bid_value') is not None):
            best[domain] = row
    return sorted(best.values(), key=lambda r: (r.get('closing_dt_utc') or '9999', r['domain']))


def main() -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    service = load_drive()
    files = list_recent_files(service)
    all_rows = []
    kept_files = []
    total_rows_scanned = 0
    total_rows_qualified = 0
    for meta in files:
        try:
            raw = download_file(service, meta)
        except Exception:
            continue
        out_name = f"{meta['id']}_{meta['name'].replace('/', '_')}"
        (RAW_DIR / out_name).write_bytes(raw)
        parsed = detect_and_parse(meta, raw)
        rows_scanned = 0
        lower = meta['name'].lower()
        try:
            if lower.endswith('.csv') or meta.get('mimeType') in {'text/csv', 'application/csv'}:
                _, raw_rows = rows_from_csv_bytes(raw)
                rows_scanned = len(raw_rows)
            elif lower.endswith('.xlsx') or meta.get('mimeType') == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                _, raw_rows = rows_from_xlsx_bytes(raw)
                rows_scanned = len(raw_rows)
        except Exception:
            rows_scanned = 0
        total_rows_scanned += rows_scanned
        total_rows_qualified += len(parsed)
        if parsed or rows_scanned:
            kept_files.append({
                'id': meta['id'],
                'name': meta['name'],
                'modifiedTime': meta.get('modifiedTime'),
                'webViewLink': meta.get('webViewLink'),
                'rowsScanned': rows_scanned,
                'rowsQualified': len(parsed),
            })
        if parsed:
            all_rows.extend(parsed)
    payload = {
        'generatedAt': datetime.now(timezone.utc).isoformat(),
        'folderId': FOLDER_ID,
        'windowHours': WINDOW_HOURS,
        'filesScanned': len(files),
        'rowsScanned': total_rows_scanned,
        'rowsQualified': total_rows_qualified,
        'files': kept_files,
        'rows': dedupe(all_rows),
    }
    OUTPUT.write_text(json.dumps(payload, indent=2) + '\n')
    print(json.dumps({'files_scanned': len(files), 'files_kept': len(kept_files), 'rows_scanned': total_rows_scanned, 'rows_qualified': total_rows_qualified, 'rows_saved': len(payload['rows']), 'output': str(OUTPUT)}, indent=2))


if __name__ == '__main__':
    main()
